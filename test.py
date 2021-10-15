import json
from openpyxl import workbook
from openpyxl import load_workbook

def main():
    #whole_dict = []
    '''
    wb = load_workbook('whole_collection_processed.xlsx')
    ws = wb['Sheet1'] 
    for row in range(5,14):
        dic = {}
        dic['part_num'] = ws['A'+str(row)].value
        dic['part_num'] = ws['B' + str(row)].value
        dic['part_url'] = ws['D' + str(row)].value
        dic['short_desc'] = ws['E' + str(row)].value
        dic['part_type'] = ws['F' + str(row)].value
        dic['team'] = ws['G' + str(row)].value
        dic['year'] = ws['H' + str(row)].value
        dic['sequence'] = ws['I' + str(row)].value
        dic['contents'] = ws['J' + str(row)].value
        dic['released'] = ws['K' + str(row)].value
        dic['sample'] = ws['L' + str(row)].value
        dic['star'] = ws['M' + str(row)].value
        dic['twins'] = ws['N' + str(row)].value
        dic['assemble'] = ws['O' + str(row)].value
        dic['part_used'] = ws['P' + str(row)].value
        dic['using_parts'] = ws['Q' + str(row)].value
        dic['len'] = ws['R' + str(row)].value
        dic['cite'] = ws['S' + str(row)].value
        dic['distri'] = ws['T' + str(row)].value
        dic['cited'] = ws['U' + str(row)].value
        dic['cited_times'] = ws['V' + str(row)].value
        whole_dict.append(dic)
    '''
    sample_return = [{'part_name' : 'BBa_TEST1', 'search_optput':['AS1','AS2','AS3','AS4'],'tree_path':'BBa_TEST1.xtml', 'related_parts': ['AR1','AR2','AR3'], 'signif': [1,2,3]}]
    with open('sample_dict_json_test.txt','w')as f:
        f.write(json.dumps(sample_return, indent=4))



main()