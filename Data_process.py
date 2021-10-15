import re
from openpyxl import workbook
from openpyxl import load_workbook
import global_vary
def get_cite_parts():
    global fpgrowth_database
    global wb, ws
    ws = global_vary.get_value('ws')
    wb = global_vary.get_value('wb')
    dic = {}
    for i in range(ws.min_row+1, ws.max_row+1):
        cites = []
        print (i)
        content = ws['J'+str(i)].value
        ID = ws['A'+str(i)].value
        split_words = content.split(' ')
        for a_word in split_words:
            if  'BBa_' in a_word:
                a_word = ''.join(filter(str.isdigit, a_word))
                a_word = 'BBa_' + a_word
                cites.append(a_word)
                if a_word not in dic:
                    dic[a_word] = [ID]
                if a_word in dic and ID not in dic[a_word]:
                    dic[a_word].append(ID)
                continue
            if re.search( "[A-Z]\d{4}" , a_word ) and not re.search("[^a-zA-Z0-9]",a_word):
                #a_word = ''.join(filter(str.isdigit, a_word))
                a_word = 'BBa_' + a_word
                cites.append(a_word)
                if a_word not in dic:
                    dic[a_word] = [ID]
                if a_word in dic and ID not in dic[a_word]:
                    dic[a_word].append(ID)
        ws['S' + str(i)].value = ' '.join(cites)
    j = 0
    for cell in ws['A']:
        row = cell.row
        value = dic.get(cell.value, [])
        if len(value):
            ws['U' + str(row)].value = ' '.join(value)
            ws['V' + str(row)].value = len(value)
            j += 1
            print(j)
    wb.save('whole_collection_processed.xlsx')
    fpgrowth_database = []
    wb = workbook.Workbook()
    ws1 = wb.active
    ws1.append(['part_num', 'cited by', "cited times"])
    for i in dic:
        ws1.append([i, str(dic[i]), len(dic[i])])
    wb.save(f'fp_db.xlsx')
    return 1

def get_signifi_score():
    wb = load_workbook('whole_collection_processed.xlsx')
    ws = wb['Sheet1']
    for cell in ws['A']:
        if cell.value == 'part_num':
            continue
        row = str(cell.row)
        cited_times = int(ws['V'+row].value)
        if ws['P'+row].value != 'None':
            used_times = len (ws['P'+row].value.split(' '))
        else:
            used_times = 0
        distri =int(ws['T'+row].value)
        sig_score = 15 * distri + 3 * used_times + 1 * cited_times
        ws['W'+row] = sig_score
    wb.save('whole_collection_processed.xlsx')
    return 1

