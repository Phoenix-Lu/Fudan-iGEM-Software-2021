from openpyxl import  load_workbook

def is_title(inputs, stdBBa):
    if (inputs.lower() in stdBBa.lower()) or (inputs.lower() in stdBBa.replace('_', '')) or (stdBBa.lower() in inputs.lower()):
        return 1
    else:
        return 0


def is_in_description(inputs, comparison):
    if inputs.lower() in comparison.lower():
        return 1
    else:
        return 0

def is_same_sequence(sequence1, sequence2):
    if sequence1.lower == sequence2:
        return 1
    reverse_sequence = sequence1.lower()
    reverse_sequence = reverse_sequence[::-1]
    reverse_sequence.replace('a', 't')
    reverse_sequence.replace('c', 'g')
    reverse_sequence.replace('t', 'a')
    reverse_sequence.replace('g', 'c')
    if reverse_sequence == sequence2:
        return 1
    else:
        return 0


def is_InRelation_seqience(sequence1, sequence2):
    if (sequence1.lower() in sequence2) or (sequence2 in sequence1.lower()):
        return 1
    reverse_sequence = sequence1.lower()
    reverse_sequence = reverse_sequence[::-1]
    reverse_sequence.replace('a', 't')
    reverse_sequence.replace('c', 'g')
    reverse_sequence.replace('t', 'a')
    reverse_sequence.replace('g', 'c')
    if reverse_sequence in sequence2 or sequence2 in reverse_sequence:
        return 1
    else:
        return 0

def take_search_result(inputs, ws):
    search_list = []
    search_dict = {'0':[], '1':[], '2':[],'3':[]}
    for cell in ws['A']:
        row = str(cell.row)
        if is_title(inputs, cell.value):
            search_dict['0'] = cell.value

        if is_same_sequence(inputs, ws['I'+row].value):
            search_dict['0'] = cell.value
        if is_InRelation_seqience(inputs, ws['I'+row].value):
            search_dict['2'].append(cell.value)
        if is_in_description(inputs, ws['B'+row].value) or is_in_description(inputs, ws['E'+row].value):
            search_dict['1'].append(cell.value)
    search_list.append(search_dict['0'])
    search_list.extend(search_dict['1'])
    search_list.extend(search_dict['2'])
    search_list.extend(search_dict['3'])
    return search_list