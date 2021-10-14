from pyecharts import options as opts
from pyecharts.charts import Tree
from openpyxl import workbook
from openpyxl import load_workbook
import global_vary

#flag暂定是如此
def switch_flag_for_kind(key, value):
    flag = -1
    if key == 'twins':
        flag = 1
    if key == 'using_parts' or key == 'used_parts':
        flag = 2
    if key == 'cites':
        flag = 3
    if get_self_distri(value):
        flag = 4
    return flag

#各个不同kind的具体参数要调
def set_opts_from_flag(a_dict):
    #if a_dict['value'] == 0:
        #a_dict['label_opts'] = {}
    if a_dict['value'] == 1:
        a_dict['label_opts'] = opts.LabelOpts(background_color = '#99ff00')
    if a_dict['value'] == 2:
        a_dict['label_opts'] = opts.LabelOpts(background_color = '#3333cc')
    if a_dict['value'] == 3:
        a_dict['label_opts'] = opts.LabelOpts()

    return a_dict


def get_next_level(part):
    global  wb, ws
    ws = global_vary.get_value('ws')
    wb = global_vary.get_value('wb')
    next_level_dict = {}
    for cell in ws['A']:
        if cell.value == part:
            row = cell.row
            if str(ws['N' + str(row)].value).split(' ')[0] != 'None':
                next_level_dict['twins'] = str(ws['N' + str(row)].value).split(' ')
            if str(ws['Q' + str(row)].value).split(' ')[0] != 'self':
                next_level_dict['using_parts'] = str(ws['Q' + str(row)].value).split(' ')
            if str(ws['P' + str(row)].value).split(' ')[0]!= 'None':
                next_level_dict['used_parts'] = str(ws['P' + str(row)].value).split(' ')
            if str(ws['S'+str(row)].value).split(' ')[0] != 'None':
                next_level_dict['cites'] = str(ws['S'+str(row)].value).split(' ')

    return next_level_dict #dict{'twins', {using_parts},{used_parts}}

def get_self_score(search_part):
    for cell in ws['A']:
        if cell.value == search_part:
            return int(ws['W'+str(cell.row)].value)

def get_self_distri(part):

    global wb, ws
    for cell in ws['A']:
        if cell.value == part:
            row = cell.row
            return int(ws['T' + str(row)].value)


def tree_data_set (central_part):
    global ws, wb
    ws = global_vary.get_value('ws')
    wb = global_vary.get_value('wb')
    central_part_bba = central_part
    central_part_score = get_self_score(central_part)

    level_zero_data = {'name' : central_part_bba, 'value' : central_part_score}
    next_level_dict = get_next_level(str(central_part_bba))
    first_children_list = []
    for key,value in next_level_dict.items():
        for item in value:
            score = get_self_score(item)
            a_dict = {'name': item, 'value': score}
            second_level_dict = get_next_level(str(item))
            second_children_list = []
            for key, value in second_level_dict.items():
                for item2 in value:
                    score = get_self_score(item2)
                    b_dict = {'name': item2, 'value': score}
                    b_dict = set_opts_from_flag(b_dict)
                    second_children_list.append(b_dict)
            a_dict['children'] = second_children_list
            first_children_list.append(a_dict)
    level_zero_data['children'] = first_children_list
    return [level_zero_data]

#饭回[[rel_lst],[sig_lst]]
def tree_data_process(tree_data):
    whole_dic = tree_data[0]
    out_put_dict = {}
    first_level_childrenlst = whole_dic['children']
    for first_level_child in first_level_childrenlst:
        key = first_level_child['name']
        value = first_level_child['value']
        out_put_dict[key] = value
        second_level_childlst = first_level_child['children']
        for second_level_child in second_level_childlst:
            key = second_level_child['name']
            value = second_level_child['value']
            out_put_dict[key] = value

    out_put_lst = [[],[]]
    for key, value in out_put_dict.items():
        out_put_lst[0].append(key)
        out_put_lst[1].append(value)
    return out_put_lst


def set_tree(data):
    c = (
        Tree(init_opts= opts.InitOpts(width= '1350px', height='750px'))#这里改界面初始,例如放大界面
        .add(
            series_name = data[0]['name'],
            data= data,
            pos_top="15%",
            pos_bottom="15%",
            pos_right='5%',
            pos_left='5%',
            layout="radial",
            symbol="emptyCircle",
            symbol_size=7,)
        .set_series_opts(label_opts = opts.LabelOpts())
        .set_global_opts(title_opts=opts.TitleOpts(title="Tree-Layout"))
        .set_global_opts(tooltip_opts=opts.TooltipOpts(trigger="item", trigger_on="mousemove"))
        .render('TreeMapResults\\' + data[0]['name'] + "tree.html")
        #可以有其他render
        )
    return 1




'''
    i, j = 0
    second_level_bba = []
    second_level_url = []

    while i <= len(first_level_bba)-1:
        dict_list = []
        a_dict = {'name' : first_level_bba[i], 'value' : first_level_url[i]  }
        dict_list.append(a_dict)
        second_level_bba.append(get_next_level(first_level_bba[i])[0])
        second_level_url.append(get_next_level(first_level_bba[i])[1])
        i += 1
    level_zero_data['children'] = dict_list
    i=0
    while i<= len(first_level_bba) -1:
        while j <= len(second_level_bba)-1:
            dict_list = []
            a_dict = {'name' : second_level_url[j], 'value' : first_level_url[j]}
            dict_list.append(a_dict)
        level_zero_data['children'][i] = [dict_list]

    return level_zero_data


'''








